import io
import os
from datetime import datetime
from functools import wraps

import pandas as pd
from flask import (Flask, flash, redirect, render_template, request,
                   send_file, url_for)
from flask_login import (LoginManager, current_user, login_required,
                         login_user, logout_user)

from models import CollectionCall, Deal, DealStatusHistory, Dealer, DealerContact, Payment, Status, User, db

# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────

class SimplePagination:
    """Pagination wrapper for in-memory sorted lists."""
    def __init__(self, items, page, per_page, total):
        self.items    = items
        self.page     = page
        self.per_page = per_page
        self.total    = total
        self.pages    = max(1, (total + per_page - 1) // per_page)
        self.has_prev = page > 1
        self.has_next = page < self.pages
        self.prev_num = page - 1
        self.next_num = page + 1

    def iter_pages(self, left_edge=1, right_edge=1, left_current=2, right_current=2):
        last = 0
        for num in range(1, self.pages + 1):
            if (num <= left_edge or
                    self.page - left_current - 1 < num < self.page + right_current or
                    num > self.pages - right_edge):
                if last + 1 != num:
                    yield None
                yield num
                last = num

STATUS_COLORS = {
    "blue":   "bg-blue-100 text-blue-800 border border-blue-200",
    "red":    "bg-red-100 text-red-800 border border-red-200",
    "yellow": "bg-yellow-100 text-yellow-800 border border-yellow-200",
    "green":  "bg-green-100 text-green-800 border border-green-200",
    "gray":   "bg-gray-100 text-gray-700 border border-gray-200",
    "purple": "bg-purple-100 text-purple-800 border border-purple-200",
    "orange": "bg-orange-100 text-orange-800 border border-orange-200",
}

STATUS_DOT = {
    "blue":   "bg-blue-500",
    "red":    "bg-red-500",
    "yellow": "bg-yellow-400",
    "green":  "bg-green-500",
    "gray":   "bg-gray-400",
    "purple": "bg-purple-500",
    "orange": "bg-orange-400",
}

COLOR_OPTIONS = ["blue", "red", "yellow", "green", "gray", "purple", "orange"]


def _duration_str(start: datetime, end: datetime) -> str:
    delta = end - start
    total_s = max(int(delta.total_seconds()), 0)
    days = total_s // 86400
    hours = (total_s % 86400) // 3600
    minutes = (total_s % 3600) // 60
    parts = []
    if days:
        parts.append(f"{days}d")
    if hours:
        parts.append(f"{hours}h")
    parts.append(f"{minutes}m")
    return " ".join(parts) or "< 1m"


def _calc_redline(system_size, company_redline, adders, contract_amount):
    total_commission = contract_amount - (system_size * company_redline) - adders
    total_ppw = contract_amount / system_size if system_size else 0.0
    net_ppw = (contract_amount - adders) / system_size if system_size else 0.0
    return round(total_commission, 2), round(total_ppw, 4), round(net_ppw, 4)


def admin_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not current_user.is_authenticated:
            flash("Please sign in to continue.", "warning")
            return redirect(url_for("login"))
        if not current_user.is_admin:
            flash("Access denied. Admin privileges required.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return wrapped


def _record_payment(deal, amount, user_id, note=None):
    """Record a payment and handle payroll status auto-transitions. Returns (paid_in_full: bool)."""
    db.session.add(Payment(
        deal_id=deal.id, amount=amount,
        recorded_by=user_id, note=note
    ))
    deal.amount_paid = round((deal.amount_paid or 0.0) + amount, 2)
    deal.updated_at = datetime.utcnow()

    # Auto-transition payroll status to Paid when balance is cleared
    if deal.remaining_balance <= 0.01:
        paid_status = Status.query.filter_by(name="Paid", type="payroll").first()
        if paid_status and deal.payroll_status_id != paid_status.id:
            deal.payroll_status_id = paid_status.id
            db.session.add(DealStatusHistory(
                deal_id=deal.id, status_id=paid_status.id,
                changed_by=user_id,
                status_type="payroll",
                note=f"Auto-transitioned: paid in full (last payment ${amount:,.2f})"
            ))
        return True

    # Auto-transition to Partial paid if not already in a payment payroll status
    payroll_status = deal.payroll_status
    if payroll_status and payroll_status.name not in ("Partial paid", "Paid"):
        partial = Status.query.filter_by(name="Partial paid", type="payroll").first()
        if partial:
            deal.payroll_status_id = partial.id
            db.session.add(DealStatusHistory(
                deal_id=deal.id, status_id=partial.id,
                changed_by=user_id,
                status_type="payroll",
                note=f"Payment of ${amount:,.2f} received"
            ))
    elif not payroll_status:
        # No payroll status set — auto-assign Partial paid
        partial = Status.query.filter_by(name="Partial paid", type="payroll").first()
        if partial:
            deal.payroll_status_id = partial.id
            db.session.add(DealStatusHistory(
                deal_id=deal.id, status_id=partial.id,
                changed_by=user_id,
                status_type="payroll",
                note=f"Payment of ${amount:,.2f} received"
            ))
    return False


# ─────────────────────────────────────────────────────────────────────────────
#  App factory
# ─────────────────────────────────────────────────────────────────────────────

def _save_contacts(dealer, form):
    names  = form.getlist("contact_name[]")
    phones = form.getlist("contact_phone[]")
    emails = form.getlist("contact_email[]")
    for name, phone, email in zip(names, phones, emails):
        name = name.strip()
        if name:
            db.session.add(DealerContact(
                dealer_id=dealer.id,
                name=name,
                phone=phone.strip() or None,
                email=email.strip() or None,
            ))


def create_app():
    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "moses-crm-secret-x9k-2024")
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///moses_crm.db"
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

    db.init_app(app)

    login_manager = LoginManager(app)
    login_manager.login_view = "login"
    login_manager.login_message = "Please sign in to continue."
    login_manager.login_message_category = "warning"

    @login_manager.user_loader
    def load_user(uid):
        return User.query.get(int(uid))

    # ── Template filters ──────────────────────────────────────────────────────
    @app.template_filter("status_badge")
    def status_badge(color):
        return STATUS_COLORS.get(color, STATUS_COLORS["gray"])

    @app.template_filter("status_dot")
    def status_dot_filter(color):
        return STATUS_DOT.get(color, STATUS_DOT["gray"])

    @app.template_filter("currency")
    def currency_filter(value):
        try:
            return f"${float(value):,.2f}"
        except (TypeError, ValueError):
            return "—"

    @app.template_filter("fmt_dt")
    def fmt_dt(value):
        if not value:
            return "—"
        return value.strftime("%b %d %Y, %H:%M")

    _MW_CHIP = {"green":"green","blue":"blue","yellow":"amber","gray":"gray",
                "red":"red","purple":"blue","orange":"amber"}
    _MW_DOT  = {"green":"#3f6651","blue":"#4f7a9a","yellow":"#c4a050",
                "gray":"#a4a895","red":"#b8533a","purple":"#7b5ea7","orange":"#c4a050"}

    @app.template_filter("mw_chip")
    def mw_chip_filter(color):
        return _MW_CHIP.get(color, "gray")

    @app.template_filter("mw_dot")
    def mw_dot_filter(color):
        return _MW_DOT.get(color, "#a4a895")

    # ── Context processor ─────────────────────────────────────────────────────
    @app.context_processor
    def inject_globals():
        ctx = {"current_dt": datetime.utcnow(), "COLOR_OPTIONS": COLOR_OPTIONS}
        if current_user.is_authenticated:
            paid = Status.query.filter_by(name="Paid", type="payroll").first()
            if paid:
                ctx["unpaid_count"] = Deal.query.filter(
                    Deal.payroll_status_id != paid.id
                ).count()
            else:
                ctx["unpaid_count"] = Deal.query.count()
        else:
            ctx["unpaid_count"] = 0
        return ctx

    # ── DB init + seed ────────────────────────────────────────────────────────
    with app.app_context():
        db.create_all()
        # Add new columns to existing tables if not present
        from sqlalchemy import text
        with db.engine.connect() as _conn:
            for _col in [
                "ALTER TABLE installers ADD COLUMN description TEXT",
                "ALTER TABLE installers ADD COLUMN website TEXT",
                "ALTER TABLE installers ADD COLUMN portal_username TEXT",
                "ALTER TABLE installers ADD COLUMN portal_password TEXT",
                "ALTER TABLE statuses ADD COLUMN type TEXT NOT NULL DEFAULT 'install'",
                "ALTER TABLE deals ADD COLUMN payroll_status_id INTEGER REFERENCES statuses(id)",
                "ALTER TABLE deal_status_history ADD COLUMN status_type TEXT NOT NULL DEFAULT 'install'",
            ]:
                try:
                    _conn.execute(text(_col))
                    _conn.commit()
                except Exception:
                    pass
            # Classify existing statuses by name
            try:
                _conn.execute(text("UPDATE statuses SET type='payroll' WHERE name IN ('Unpaid','Partial paid','Paid')"))
                _conn.execute(text("UPDATE statuses SET type='install' WHERE name IN ('NTP','On going')"))
                _conn.commit()
            except Exception:
                pass
            # For deals whose current_status_id points to a payroll status:
            # move it to payroll_status_id and clear current_status_id
            try:
                _conn.execute(text("""
                    UPDATE deals SET
                        payroll_status_id = current_status_id,
                        current_status_id = NULL
                    WHERE current_status_id IN (
                        SELECT id FROM statuses WHERE type = 'payroll'
                    ) AND payroll_status_id IS NULL
                """))
                _conn.commit()
            except Exception:
                pass

        # Ensure every deal has both an install and payroll status
        _on_going = Status.query.filter_by(name="On going", type="install").first()
        _unpaid   = Status.query.filter_by(name="Unpaid",   type="payroll").first()
        if _on_going or _unpaid:
            for _d in Deal.query.all():
                if _d.current_status_id is None and _on_going:
                    _d.current_status_id = _on_going.id
                if _d.payroll_status_id is None and _unpaid:
                    _d.payroll_status_id = _unpaid.id
            db.session.commit()

        if not User.query.filter_by(email="Admin@harpaudit.com").first():
            admin = User(email="Admin@harpaudit.com", role="admin")
            admin.set_password("Orion123#")
            db.session.add(admin)
        if Status.query.count() == 0:
            for name, order, color, stype in [
                ("NTP",          1, "yellow", "install"),
                ("On going",     2, "blue",   "install"),
                ("Unpaid",       1, "red",    "payroll"),
                ("Partial paid", 2, "yellow", "payroll"),
                ("Paid",         3, "green",  "payroll"),
            ]:
                db.session.add(Status(name=name, order=order, color=color, is_default=True, type=stype))
        db.session.commit()

    # ═════════════════════════════════════════════════════════════════════════
    #  AUTH
    # ═════════════════════════════════════════════════════════════════════════

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for("dashboard"))
        if request.method == "POST":
            email = request.form.get("email", "").strip()
            password = request.form.get("password", "")
            user = User.query.filter_by(email=email).first()
            if user and user.check_password(password):
                login_user(user, remember=True)
                return redirect(request.args.get("next") or url_for("dashboard"))
            flash("Invalid email or password.", "error")
        all_d = Deal.query.all()
        return render_template(
            "login.html",
            login_deals=Deal.query.count(),
            login_dealers=Dealer.query.filter_by(is_active=True).count(),
            login_pipeline=sum(d.pipeline_value for d in all_d),
            login_collected=sum((d.amount_paid or 0.0) for d in all_d),
        )

    @app.route("/logout")
    @login_required
    def logout():
        logout_user()
        flash("You have been signed out.", "info")
        return redirect(url_for("login"))

    # ═════════════════════════════════════════════════════════════════════════
    #  DASHBOARD
    # ═════════════════════════════════════════════════════════════════════════

    @app.route("/")
    @login_required
    def dashboard():
        installer_filter  = request.args.get("installer_id", "")
        # install_filter accepts both new "install_id" param and legacy "status_id" param
        install_filter    = request.args.get("install_id", "") or request.args.get("status_id", "")
        payroll_filter    = request.args.get("payroll_id", "")
        page              = request.args.get("page", 1, type=int)
        per_page_raw      = request.args.get("per_page", 10, type=int)
        per_page          = per_page_raw if per_page_raw in (10, 20, 50) else 10
        sort              = request.args.get("sort", "created")
        direction         = request.args.get("dir", "desc")
        if sort not in ("created", "original", "remaining"):
            sort = "created"
        if direction not in ("asc", "desc"):
            direction = "desc"

        query = Deal.query
        if installer_filter:
            query = query.filter_by(installer_id=int(installer_filter))
        if install_filter:
            query = query.filter_by(current_status_id=int(install_filter))
        if payroll_filter:
            query = query.filter_by(payroll_status_id=int(payroll_filter))

        if sort == "created":
            order_col = Deal.created_at.desc() if direction == "desc" else Deal.created_at.asc()
            pagination = query.order_by(order_col).paginate(page=page, per_page=per_page, error_out=False)
        else:
            key_fn = (lambda d: d.original_value) if sort == "original" else (lambda d: d.remaining_balance)
            all_filtered = sorted(query.all(), key=key_fn, reverse=(direction == "desc"))
            total  = len(all_filtered)
            start  = (page - 1) * per_page
            pagination = SimplePagination(all_filtered[start:start + per_page], page, per_page, total)

        deals = pagination.items

        all_deals        = Deal.query.all()
        dealers          = Dealer.query.filter_by(is_active=True).order_by(Dealer.name).all()
        install_statuses = Status.query.filter_by(type='install').order_by(Status.order).all()
        payroll_statuses = Status.query.filter_by(type='payroll').order_by(Status.order).all()
        # Combined list for legacy template references
        statuses         = install_statuses + payroll_statuses

        paid_status    = Status.query.filter_by(name="Paid", type="payroll").first()
        total_pipeline = sum(d.pipeline_value for d in all_deals)

        total_collected = 0.0
        for d in all_deals:
            if (d.amount_paid or 0) > 0:
                total_collected += d.amount_paid
            elif paid_status and d.payroll_status_id == paid_status.id:
                total_collected += d.original_value

        dealer_kpis = [
            {
                "id": inst.id,
                "name": inst.name,
                "count": sum(1 for d in all_deals if d.installer_id == inst.id),
                "total": sum(d.pipeline_value for d in all_deals if d.installer_id == inst.id),
                "collected": sum(
                    (d.amount_paid if (d.amount_paid or 0) > 0
                     else (d.original_value if paid_status and d.payroll_status_id == paid_status.id else 0))
                    for d in all_deals if d.installer_id == inst.id
                ),
            }
            for inst in dealers
            if any(d.installer_id == inst.id for d in all_deals)
        ]
        dealer_kpis.sort(key=lambda x: x["total"], reverse=True)

        install_kpis = [
            {
                "id": s.id,
                "name": s.name,
                "color": s.color,
                "count": sum(1 for d in all_deals if d.current_status_id == s.id),
            }
            for s in install_statuses
        ]
        payroll_kpis = [
            {
                "id": s.id,
                "name": s.name,
                "color": s.color,
                "count": sum(1 for d in all_deals if d.payroll_status_id == s.id),
            }
            for s in payroll_statuses
        ]
        status_kpis = install_kpis + payroll_kpis

        return render_template(
            "dashboard.html",
            deals=deals,
            pagination=pagination,
            per_page=per_page,
            sort=sort,
            direction=direction,
            all_deals=all_deals,
            dealers=dealers,
            statuses=statuses,
            install_statuses=install_statuses,
            payroll_statuses=payroll_statuses,
            total_deals=len(all_deals),
            total_pipeline=total_pipeline,
            total_collected=total_collected,
            dealer_kpis=dealer_kpis,
            status_kpis=status_kpis,
            install_kpis=install_kpis,
            payroll_kpis=payroll_kpis,
            installer_filter=installer_filter,
            install_filter=install_filter,
            payroll_filter=payroll_filter,
            # legacy alias
            status_filter=install_filter,
        )

    # ═════════════════════════════════════════════════════════════════════════
    #  DEALS – CREATE
    # ═════════════════════════════════════════════════════════════════════════

    @app.route("/deals/new", methods=["GET", "POST"])
    @login_required
    def deal_new():
        dealers          = Dealer.query.filter_by(is_active=True).order_by(Dealer.name).all()
        install_statuses = Status.query.filter_by(type='install').order_by(Status.order).all()
        payroll_statuses = Status.query.filter_by(type='payroll').order_by(Status.order).all()
        statuses         = install_statuses + payroll_statuses
        default_install  = install_statuses[0] if install_statuses else None
        default_payroll  = payroll_statuses[0] if payroll_statuses else None

        if request.method == "POST":
            name         = request.form.get("name", "").strip()
            installer_id = request.form.get("installer_id", type=int)
            use_redline  = request.form.get("use_redline") == "1"
            notes        = request.form.get("notes", "").strip()
            status_id    = request.form.get("status_id", type=int) or (
                default_install.id if default_install else None
            )
            payroll_status_id = request.form.get("payroll_status_id", type=int)
            if not payroll_status_id:
                payroll_status_id = default_payroll.id if default_payroll else None

            if not name or not installer_id:
                if request.form.get("_modal") == "1":
                    from flask import jsonify
                    return jsonify({"ok": False, "error": "Deal name and dealer are required."}), 400
                flash("Deal name and installer are required.", "error")
                return render_template("deal_form.html", deal=None,
                                       dealers=dealers, statuses=statuses,
                                       install_statuses=install_statuses,
                                       payroll_statuses=payroll_statuses)

            deal = Deal(
                name=name, installer_id=installer_id, use_redline=use_redline,
                notes=notes, current_status_id=status_id,
                payroll_status_id=payroll_status_id,
                created_by=current_user.id,
            )

            if use_redline:
                try:
                    ss = float(request.form.get("system_size") or 0)
                    cr = float(request.form.get("company_redline") or 0)
                    ad = float(request.form.get("adders") or 0)
                    ca = float(request.form.get("contract_amount") or 0)
                except ValueError:
                    flash("Financial fields must be valid numbers.", "error")
                    return render_template("deal_form.html", deal=None,
                                           dealers=dealers, statuses=statuses,
                                           install_statuses=install_statuses,
                                           payroll_statuses=payroll_statuses)
                deal.system_size = ss
                deal.company_redline = cr
                deal.adders = ad
                deal.contract_amount = ca
                deal.total_commission, deal.total_ppw, deal.net_ppw = _calc_redline(ss, cr, ad, ca)
            else:
                try:
                    deal.amount_owed = float(request.form.get("amount_owed") or 0)
                except ValueError:
                    flash("Amount must be a valid number.", "error")
                    return render_template("deal_form.html", deal=None,
                                           dealers=dealers, statuses=statuses,
                                           install_statuses=install_statuses,
                                           payroll_statuses=payroll_statuses)

            db.session.add(deal)
            db.session.flush()

            if status_id:
                db.session.add(DealStatusHistory(
                    deal_id=deal.id, status_id=status_id, changed_by=current_user.id,
                    status_type="install"
                ))
            if payroll_status_id:
                db.session.add(DealStatusHistory(
                    deal_id=deal.id, status_id=payroll_status_id, changed_by=current_user.id,
                    status_type="payroll"
                ))

            db.session.commit()

            if request.form.get("_modal") == "1":
                from flask import jsonify
                return jsonify({"ok": True, "deal_id": deal.id, "deal_name": name})

            flash(f'Deal "{name}" created successfully.', "success")
            return redirect(url_for("deal_detail", deal_id=deal.id))

        return render_template("deal_form.html", deal=None,
                               dealers=dealers, statuses=statuses,
                               install_statuses=install_statuses,
                               payroll_statuses=payroll_statuses)

    # ═════════════════════════════════════════════════════════════════════════
    #  DEALS – DETAIL
    # ═════════════════════════════════════════════════════════════════════════

    @app.route("/deals/<int:deal_id>")
    @login_required
    def deal_detail(deal_id):
        deal             = Deal.query.get_or_404(deal_id)
        install_statuses = Status.query.filter_by(type='install').order_by(Status.order).all()
        payroll_statuses = Status.query.filter_by(type='payroll').order_by(Status.order).all()
        statuses         = install_statuses + payroll_statuses
        dealers          = Dealer.query.filter_by(is_active=True).order_by(Dealer.name).all()
        history          = deal.status_history

        history_enriched = []
        for i, entry in enumerate(history):
            end_time = history[i + 1].changed_at if i + 1 < len(history) else datetime.utcnow()
            history_enriched.append({
                "entry": entry,
                "duration": _duration_str(entry.changed_at, end_time),
                "is_current": i + 1 == len(history),
            })

        return render_template("deal_detail.html", deal=deal,
                               statuses=statuses,
                               install_statuses=install_statuses,
                               payroll_statuses=payroll_statuses,
                               dealers=dealers,
                               history=history_enriched)

    # ═════════════════════════════════════════════════════════════════════════
    #  DEALS – EDIT
    # ═════════════════════════════════════════════════════════════════════════

    @app.route("/deals/<int:deal_id>/edit", methods=["GET", "POST"])
    @login_required
    def deal_edit(deal_id):
        deal             = Deal.query.get_or_404(deal_id)
        dealers          = Dealer.query.filter_by(is_active=True).order_by(Dealer.name).all()
        install_statuses = Status.query.filter_by(type='install').order_by(Status.order).all()
        payroll_statuses = Status.query.filter_by(type='payroll').order_by(Status.order).all()
        statuses         = install_statuses + payroll_statuses

        if request.method == "POST":
            deal.name         = request.form.get("name", "").strip()
            deal.installer_id = request.form.get("installer_id", type=int)
            deal.use_redline  = request.form.get("use_redline") == "1"
            deal.notes        = request.form.get("notes", "").strip()
            new_status_id         = request.form.get("status_id", type=int)
            new_payroll_status_id = request.form.get("payroll_status_id", type=int)

            if deal.use_redline:
                try:
                    ss = float(request.form.get("system_size") or 0)
                    cr = float(request.form.get("company_redline") or 0)
                    ad = float(request.form.get("adders") or 0)
                    ca = float(request.form.get("contract_amount") or 0)
                except ValueError:
                    flash("Invalid numeric values.", "error")
                    return render_template("deal_form.html", deal=deal,
                                           dealers=dealers, statuses=statuses,
                                           install_statuses=install_statuses,
                                           payroll_statuses=payroll_statuses)
                deal.system_size, deal.company_redline = ss, cr
                deal.adders, deal.contract_amount = ad, ca
                deal.total_commission, deal.total_ppw, deal.net_ppw = _calc_redline(ss, cr, ad, ca)
                deal.amount_owed = None
            else:
                try:
                    deal.amount_owed = float(request.form.get("amount_owed") or 0)
                except ValueError:
                    flash("Invalid amount.", "error")
                    return render_template("deal_form.html", deal=deal,
                                           dealers=dealers, statuses=statuses,
                                           install_statuses=install_statuses,
                                           payroll_statuses=payroll_statuses)
                deal.system_size = deal.company_redline = deal.adders = None
                deal.contract_amount = deal.total_commission = deal.total_ppw = deal.net_ppw = None

            if new_status_id and new_status_id != deal.current_status_id:
                deal.current_status_id = new_status_id
                db.session.add(DealStatusHistory(
                    deal_id=deal.id, status_id=new_status_id, changed_by=current_user.id,
                    status_type="install"
                ))

            if new_payroll_status_id and new_payroll_status_id != deal.payroll_status_id:
                deal.payroll_status_id = new_payroll_status_id
                db.session.add(DealStatusHistory(
                    deal_id=deal.id, status_id=new_payroll_status_id, changed_by=current_user.id,
                    status_type="payroll"
                ))

            deal.updated_at = datetime.utcnow()
            db.session.commit()

            if request.form.get("_modal") == "1":
                from flask import jsonify
                return jsonify({"ok": True, "deal_name": deal.name})

            flash("Deal updated successfully.", "success")
            return redirect(url_for("deal_detail", deal_id=deal.id))

        return render_template("deal_form.html", deal=deal,
                               dealers=dealers, statuses=statuses,
                               install_statuses=install_statuses,
                               payroll_statuses=payroll_statuses)

    @app.route("/deals/<int:deal_id>/data")
    @login_required
    def deal_data(deal_id):
        from flask import jsonify
        d = Deal.query.get_or_404(deal_id)
        return jsonify({
            "id":                 d.id,
            "name":               d.name,
            "installer_id":       d.installer_id,
            "current_status_id":  d.current_status_id,
            "payroll_status_id":  d.payroll_status_id,
            "notes":              d.notes or "",
            "use_redline":        d.use_redline,
            "system_size":        d.system_size,
            "company_redline":    d.company_redline,
            "adders":             d.adders,
            "contract_amount":    d.contract_amount,
            "amount_owed":        d.amount_owed,
        })

    # ═════════════════════════════════════════════════════════════════════════
    #  DEALS – UPDATE STATUS
    # ═════════════════════════════════════════════════════════════════════════

    @app.route("/deals/<int:deal_id>/update_status", methods=["POST"])
    @login_required
    def deal_update_status(deal_id):
        deal          = Deal.query.get_or_404(deal_id)
        status_type   = request.form.get("status_type", "install")  # 'install' or 'payroll'
        new_status_id = request.form.get("status_id", type=int)
        note          = request.form.get("note", "").strip() or None

        if status_type == "payroll":
            current_id = deal.payroll_status_id
            if new_status_id and new_status_id != current_id:
                deal.payroll_status_id = new_status_id
                deal.updated_at = datetime.utcnow()
                db.session.add(DealStatusHistory(
                    deal_id=deal.id, status_id=new_status_id,
                    changed_by=current_user.id, note=note,
                    status_type="payroll"
                ))
                db.session.commit()
                flash("Payroll status updated.", "success")
            elif note and current_id:
                db.session.add(DealStatusHistory(
                    deal_id=deal.id, status_id=current_id,
                    changed_by=current_user.id, note=note,
                    status_type="payroll"
                ))
                db.session.commit()
                flash("Note added.", "success")
            else:
                flash("Select a different status or add a note.", "warning")
        else:
            current_id = deal.current_status_id
            if new_status_id and new_status_id != current_id:
                deal.current_status_id = new_status_id
                deal.updated_at = datetime.utcnow()
                db.session.add(DealStatusHistory(
                    deal_id=deal.id, status_id=new_status_id,
                    changed_by=current_user.id, note=note,
                    status_type="install"
                ))
                db.session.commit()
                flash("Install status updated.", "success")
            elif note and current_id:
                db.session.add(DealStatusHistory(
                    deal_id=deal.id, status_id=current_id,
                    changed_by=current_user.id, note=note,
                    status_type="install"
                ))
                db.session.commit()
                flash("Note added.", "success")
            else:
                flash("Select a different status or add a note.", "warning")

        return redirect(url_for("deal_detail", deal_id=deal_id))

    # ═════════════════════════════════════════════════════════════════════════
    #  DEALS – ADD PAYMENT (additional payments on existing partial deals)
    # ═════════════════════════════════════════════════════════════════════════

    @app.route("/deals/<int:deal_id>/add_payment", methods=["POST"])
    @login_required
    def deal_add_payment(deal_id):
        deal = Deal.query.get_or_404(deal_id)
        note = request.form.get("note", "").strip() or None

        try:
            amount = float(request.form.get("payment_amount") or 0)
        except ValueError:
            flash("Invalid payment amount.", "error")
            return redirect(url_for("deal_detail", deal_id=deal_id))

        if amount <= 0:
            flash("Payment amount must be greater than 0.", "error")
            return redirect(url_for("deal_detail", deal_id=deal_id))

        if amount > deal.remaining_balance + 0.01:
            flash(f"Payment (${amount:,.2f}) exceeds remaining balance (${deal.remaining_balance:,.2f}).", "error")
            return redirect(url_for("deal_detail", deal_id=deal_id))

        paid_in_full = _record_payment(deal, amount, current_user.id, note)
        db.session.commit()

        if paid_in_full:
            flash(f"Payment of ${amount:,.2f} recorded — deal automatically marked as Paid!", "success")
        else:
            flash(f"Payment of ${amount:,.2f} recorded. Remaining: ${deal.remaining_balance:,.2f}", "success")

        return redirect(url_for("deal_detail", deal_id=deal_id))

    # ═════════════════════════════════════════════════════════════════════════
    #  PAYMENTS – DELETE
    # ═════════════════════════════════════════════════════════════════════════

    @app.route("/deals/<int:deal_id>/payments/<int:payment_id>/delete", methods=["POST"])
    @login_required
    @admin_required
    def payment_delete(deal_id, payment_id):
        deal = Deal.query.get_or_404(deal_id)
        payment = Payment.query.filter_by(id=payment_id, deal_id=deal_id).first_or_404()

        deleted_amount = payment.amount
        db.session.delete(payment)
        db.session.flush()

        # Recalculate from remaining payments to avoid float drift
        deal.amount_paid = round(sum(p.amount for p in deal.payments), 2)
        deal.updated_at = datetime.utcnow()

        # Auto-adjust payroll status after deletion
        new_payroll_status = None
        if deal.amount_paid <= 0:
            # No payments left — revert payroll to Unpaid
            new_payroll_status = Status.query.filter_by(name="Unpaid", type="payroll").first()
            if not new_payroll_status:
                # Fallback: first payroll status
                new_payroll_status = Status.query.filter_by(type="payroll").order_by(Status.order).first()
        elif deal.remaining_balance > 0.01 and deal.payroll_status and deal.payroll_status.name == "Paid":
            # Was fully paid but now has remaining balance — downgrade to Partial paid
            new_payroll_status = Status.query.filter_by(name="Partial paid", type="payroll").first()

        if new_payroll_status and deal.payroll_status_id != new_payroll_status.id:
            deal.payroll_status_id = new_payroll_status.id
            db.session.add(DealStatusHistory(
                deal_id=deal.id, status_id=new_payroll_status.id,
                changed_by=current_user.id,
                status_type="payroll",
                note=f"Auto-adjusted: payment of ${deleted_amount:,.2f} deleted"
            ))

        db.session.commit()
        flash(f"Payment of ${deleted_amount:,.2f} deleted.", "success")
        return redirect(url_for("deal_detail", deal_id=deal_id))

    # ═════════════════════════════════════════════════════════════════════════
    #  DEALS – DELETE
    # ═════════════════════════════════════════════════════════════════════════

    @app.route("/deals/<int:deal_id>/delete", methods=["POST"])
    @login_required
    @admin_required
    def deal_delete(deal_id):
        deal = Deal.query.get_or_404(deal_id)
        Payment.query.filter_by(deal_id=deal_id).delete()
        DealStatusHistory.query.filter_by(deal_id=deal_id).delete()
        db.session.delete(deal)
        db.session.commit()
        flash("Deal deleted.", "success")
        return redirect(url_for("dashboard"))

    # ═════════════════════════════════════════════════════════════════════════
    #  DEALERS
    # ═════════════════════════════════════════════════════════════════════════

    @app.route("/dealers/<int:dealer_id>/detail")
    @login_required
    def dealer_detail(dealer_id):
        dealer = Dealer.query.get_or_404(dealer_id)
        deals = (Deal.query
                 .filter_by(installer_id=dealer_id)
                 .order_by(Deal.created_at.desc())
                 .all())
        collection_calls = (CollectionCall.query
                            .filter_by(dealer_id=dealer_id)
                            .order_by(CollectionCall.created_at.desc())
                            .all())
        return render_template("dealer_detail.html",
                               dealer=dealer,
                               deals=deals,
                               collection_calls=collection_calls)

    @app.route("/dealers/<int:dealer_id>/collection_calls/add", methods=["POST"])
    @login_required
    def dealer_add_collection_call(dealer_id):
        dealer = Dealer.query.get_or_404(dealer_id)
        body = request.form.get("body", "").strip()
        if not body:
            flash("Note cannot be empty.", "error")
        else:
            db.session.add(CollectionCall(
                dealer_id=dealer.id,
                body=body,
                created_by=current_user.id,
            ))
            db.session.commit()
            flash("Note added.", "success")
        return redirect(url_for("dealer_detail", dealer_id=dealer_id))

    @app.route("/dealers")
    @login_required
    def dealers_list():
        dealers = Dealer.query.order_by(Dealer.is_active.desc(), Dealer.name).all()
        return render_template("dealers.html", dealers=dealers)

    @app.route("/dealers/new", methods=["GET", "POST"])
    @login_required
    @admin_required
    def dealer_new():
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            if not name:
                flash("Dealer name is required.", "error")
            elif Dealer.query.filter_by(name=name).first():
                flash("A dealer with that name already exists.", "error")
            else:
                dealer = Dealer(
                    name=name,
                    description=request.form.get("description", "").strip() or None,
                    website=request.form.get("website", "").strip() or None,
                    portal_username=request.form.get("portal_username", "").strip() or None,
                    portal_password=request.form.get("portal_password", "").strip() or None,
                )
                db.session.add(dealer)
                db.session.flush()
                _save_contacts(dealer, request.form)
                db.session.commit()
                flash(f'Dealer "{name}" created.', "success")
                return redirect(url_for("dealers_list"))
        return render_template("dealer_form.html", dealer=None)

    @app.route("/dealers/<int:dealer_id>/edit", methods=["GET", "POST"])
    @login_required
    @admin_required
    def dealer_edit(dealer_id):
        dealer = Dealer.query.get_or_404(dealer_id)
        if request.method == "POST":
            new_name = request.form.get("name", "").strip()
            conflict = Dealer.query.filter(
                Dealer.name == new_name, Dealer.id != dealer_id
            ).first()
            if not new_name:
                flash("Dealer name is required.", "error")
            elif conflict:
                flash("That name is already taken by another dealer.", "error")
            else:
                dealer.name = new_name
                dealer.description = request.form.get("description", "").strip() or None
                dealer.website = request.form.get("website", "").strip() or None
                dealer.portal_username = request.form.get("portal_username", "").strip() or None
                dealer.portal_password = request.form.get("portal_password", "").strip() or None
                dealer.is_active = request.form.get("is_active") == "1"
                # Replace contacts
                for c in list(dealer.contacts):
                    db.session.delete(c)
                db.session.flush()
                _save_contacts(dealer, request.form)
                db.session.commit()
                flash("Dealer updated.", "success")
                return redirect(url_for("dealers_list"))
        return render_template("dealer_form.html", dealer=dealer)

    @app.route("/dealers/<int:dealer_id>/delete", methods=["POST"])
    @login_required
    @admin_required
    def dealer_delete(dealer_id):
        dealer = Dealer.query.get_or_404(dealer_id)
        if dealer.deals:
            dealer.is_active = False
            db.session.commit()
            flash("Dealer deactivated (has associated deals).", "warning")
        else:
            db.session.delete(dealer)
            db.session.commit()
            flash("Dealer deleted.", "success")
        return redirect(url_for("dealers_list"))

    # ═════════════════════════════════════════════════════════════════════════
    #  ADMIN – STATUSES
    # ═════════════════════════════════════════════════════════════════════════

    @app.route("/admin/statuses", methods=["GET", "POST"])
    @login_required
    @admin_required
    def admin_statuses():
        if request.method == "POST":
            action = request.form.get("action")

            if action == "create":
                name  = request.form.get("name", "").strip()
                order = request.form.get("order", 99, type=int)
                color = request.form.get("color", "gray")
                stype = request.form.get("type", "install")
                if stype not in ("install", "payroll"):
                    stype = "install"
                if not name:
                    flash("Status name is required.", "error")
                elif Status.query.filter_by(name=name).first():
                    flash("A status with that name already exists.", "error")
                else:
                    # Shift down statuses of the same type at or after the target position
                    for st in Status.query.filter(Status.type == stype, Status.order >= order).all():
                        st.order += 1
                    db.session.add(Status(name=name, order=order, color=color, type=stype))
                    db.session.commit()
                    flash(f'Status "{name}" created.', "success")

            elif action == "update":
                s = Status.query.get_or_404(request.form.get("status_id", type=int))
                new_order = request.form.get("order", s.order, type=int)
                if new_order != s.order:
                    old_order = s.order
                    if new_order < old_order:
                        for st in Status.query.filter(Status.order >= new_order, Status.order < old_order, Status.id != s.id).all():
                            st.order += 1
                    else:
                        for st in Status.query.filter(Status.order > old_order, Status.order <= new_order, Status.id != s.id).all():
                            st.order -= 1
                s.name  = request.form.get("name", s.name).strip()
                s.order = new_order
                s.color = request.form.get("color", s.color)
                db.session.commit()
                flash("Status updated.", "success")

            elif action == "delete":
                s = Status.query.get_or_404(request.form.get("status_id", type=int))
                has_install_deals = bool(s.deals)
                has_payroll_deals = bool(s.payroll_deals)
                if has_install_deals or has_payroll_deals:
                    flash("Cannot delete: deals are using this status.", "error")
                else:
                    db.session.delete(s)
                    db.session.commit()
                    flash("Status deleted.", "success")

            return redirect(url_for("admin_statuses"))

        install_statuses = Status.query.filter_by(type='install').order_by(Status.order).all()
        payroll_statuses = Status.query.filter_by(type='payroll').order_by(Status.order).all()
        statuses = install_statuses + payroll_statuses
        return render_template("admin_statuses.html",
                               statuses=statuses,
                               install_statuses=install_statuses,
                               payroll_statuses=payroll_statuses)

    @app.route("/admin/statuses/reorder", methods=["POST"])
    @login_required
    @admin_required
    def admin_statuses_reorder():
        items = request.get_json() or []
        for item in items:
            s = Status.query.get(item["id"])
            if s:
                s.order = item["order"]
        db.session.commit()
        return {"ok": True}

    # ═════════════════════════════════════════════════════════════════════════
    #  EXPORT
    # ═════════════════════════════════════════════════════════════════════════

    @app.route("/export")
    @login_required
    def export_excel():
        installer_filter = request.args.get("installer_id", "")
        status_filter = request.args.get("status_id", "")

        query = Deal.query
        if installer_filter:
            query = query.filter_by(installer_id=int(installer_filter))
        if status_filter:
            query = query.filter_by(current_status_id=int(status_filter))
        deals = query.order_by(Deal.created_at.desc()).all()

        rows = []
        for d in deals:
            rows.append({
                "ID": d.id,
                "Deal Name": d.name,
                "Dealer": d.dealer.name if d.dealer else "",
                "Status": d.current_status.name if d.current_status else "",
                "Calculation Type": "Redline" if d.use_redline else "Fixed Amount",
                "System Size (W)": d.system_size or "",
                "Company Redline ($/W)": d.company_redline or "",
                "Adders ($)": d.adders or "",
                "Contract Amount ($)": d.contract_amount or "",
                "Total Commission ($)": d.total_commission or "",
                "Total PPW ($/W)": d.total_ppw or "",
                "Net PPW ($/W)": d.net_ppw or "",
                "Amount Owed ($)": d.amount_owed or "",
                "Original Value ($)": d.original_value,
                "Amount Paid ($)": d.amount_paid or 0,
                "Remaining Balance ($)": d.remaining_balance,
                "Notes": d.notes or "",
                "Created": d.created_at.strftime("%Y-%m-%d %H:%M") if d.created_at else "",
                "Last Updated": d.updated_at.strftime("%Y-%m-%d %H:%M") if d.updated_at else "",
            })

        df = pd.DataFrame(rows)
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Deals")
            ws = writer.sheets["Deals"]
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        output.seek(0)
        filename = f"MOSES_deals_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    return app


# ─────────────────────────────────────────────────────────────────────────────
app = create_app()

if __name__ == "__main__":
    # Port 5000 is taken by AirPlay on macOS
    app.run(debug=True, host="0.0.0.0", port=5001)
